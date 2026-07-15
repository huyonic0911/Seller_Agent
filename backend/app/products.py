"""ProductStore: nạp danh mục sản phẩm từ JSON hoặc Excel và cung cấp truy vấn.

Ở MVP toàn bộ danh mục được nạp vào system prompt của Claude (kèm prompt caching),
nên `catalog_text()` là hàm quan trọng nhất. `search()` để sẵn cho việc nâng cấp
RAG khi danh mục lớn.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any


class ProductStore:
    def __init__(self, path: Path):
        self.path = Path(path)
        self.shop: dict[str, Any] = {}
        self.products: list[dict[str, Any]] = []
        self.load()

    # ---- Nạp dữ liệu ------------------------------------------------------
    def load(self) -> None:
        suffix = self.path.suffix.lower()
        if suffix == ".json":
            self._load_json()
        elif suffix in {".xlsx", ".xls"}:
            self._load_excel()
        else:
            raise ValueError(f"Định dạng danh mục không hỗ trợ: {suffix}")

    def _load_json(self) -> None:
        data = json.loads(self.path.read_text(encoding="utf-8"))
        self.shop = data.get("shop", {})
        self.products = data.get("products", [])

    def _load_excel(self) -> None:
        # Mỗi dòng = 1 sản phẩm; cột: id, ten, gia, gia_km, mau, size, ton_kho, mo_ta
        from openpyxl import load_workbook

        wb = load_workbook(self.path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            self.products = []
            return
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]

        def split_multi(val: Any) -> list[str]:
            if val is None:
                return []
            return [x.strip() for x in str(val).replace(";", ",").split(",") if x.strip()]

        products: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = dict(zip(headers, row))
            if not record.get("id"):
                continue
            products.append(
                {
                    "id": record.get("id"),
                    "ten": record.get("ten"),
                    "gia": record.get("gia"),
                    "gia_km": record.get("gia_km"),
                    "mau": split_multi(record.get("mau")),
                    "size": split_multi(record.get("size")),
                    "ton_kho": record.get("ton_kho"),
                    "mo_ta": record.get("mo_ta"),
                    "faq": split_multi(record.get("faq")),
                }
            )
        self.products = products
        self.shop = {}

    # ---- Truy vấn ---------------------------------------------------------
    def all(self) -> list[dict[str, Any]]:
        return self.products

    def search(self, query: str, limit: int = 5) -> list[dict[str, Any]]:
        """Tìm kiếm từ khóa đơn giản (đủ cho MVP nếu cần lọc trước khi đưa vào prompt)."""
        q = (query or "").lower()
        if not q:
            return self.products[:limit]
        scored: list[tuple[int, dict[str, Any]]] = []
        for p in self.products:
            haystack = " ".join(
                str(x).lower()
                for x in [p.get("ten"), p.get("mo_ta"), p.get("id"), " ".join(p.get("mau", []))]
            )
            score = sum(1 for token in q.split() if token in haystack)
            if score:
                scored.append((score, p))
        scored.sort(key=lambda t: t[0], reverse=True)
        return [p for _, p in scored[:limit]] or self.products[:limit]

    # ---- Dựng context cho LLM --------------------------------------------
    @staticmethod
    def _fmt_price(v: Any) -> str:
        try:
            return f"{int(v):,}đ".replace(",", ".")
        except (TypeError, ValueError):
            return str(v)

    def _product_line(self, p: dict[str, Any]) -> str:
        gia = self._fmt_price(p.get("gia"))
        parts = [f"[{p.get('id')}] {p.get('ten')}", f"giá {gia}"]
        if p.get("gia_km"):
            parts.append(f"KHUYẾN MÃI còn {self._fmt_price(p['gia_km'])}")
        ton = p.get("ton_kho")
        if ton == 0:
            parts.append("HẾT HÀNG")
        elif ton is not None:
            parts.append(f"tồn {ton}")
        if p.get("mau"):
            parts.append("màu: " + ", ".join(p["mau"]))
        if p.get("size"):
            parts.append("size: " + ", ".join(str(s) for s in p["size"]))
        if p.get("mo_ta"):
            parts.append("mô tả: " + str(p["mo_ta"]))
        if p.get("faq"):
            parts.append("FAQ: " + " | ".join(p["faq"]))
        return " — ".join(parts)

    def catalog_text(self) -> str:
        """Chuỗi mô tả toàn bộ danh mục + chính sách shop, đưa vào system prompt."""
        lines: list[str] = []
        shop_name = self.shop.get("name")
        if shop_name:
            lines.append(f"TÊN SHOP: {shop_name}")
        policies = self.shop.get("policies") or {}
        if policies:
            lines.append("CHÍNH SÁCH SHOP:")
            for k, v in policies.items():
                lines.append(f"- {k}: {v}")
        lines.append("")
        lines.append("DANH MỤC SẢN PHẨM:")
        for p in self.products:
            lines.append("- " + self._product_line(p))
        return "\n".join(lines)

    def shop_name(self) -> str:
        return self.shop.get("name", "shop mình")
